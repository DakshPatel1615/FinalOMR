from openpyxl import Workbook, load_workbook
from tkinter import *
import tkinter.messagebox


def addHeader(ws):
    ws.append(['NAME', 'CLASS', 'STREAM', 'SUBJECT', 'SCORE', 'PERCENTAGE'])


def addStudentRecord(ws, name, nameOfClass, stream, subject, score, percentage):
    ws.append([name, nameOfClass, stream, subject, score, str(percentage) + '%'])


def getWorkbook(classFile, fileList, savePath):
    if classFile in fileList:
        wb = load_workbook(savePath)
    else:
        wb = Workbook()
    return wb


def getWorksheet(exam, wb):
    if exam not in wb.sheetnames:
        wb.create_sheet(exam)
        ws = wb[exam]
        addHeader(ws)
    else:
        tkinter.messagebox.showinfo("Error Uploading",
                                    "The data for this exam has already been uploaded for this class! Try Again! T_T")
    return ws
